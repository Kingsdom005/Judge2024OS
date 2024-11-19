import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 准备数据
data = [
    {'id': '2024140224', 'name': '李华', 'score': 10},
    {'id': '2024140924', 'name': '李华2', 'score': 8},
    {'id': '2024140260', 'name': '2李华', 'score': 5}
]

def generate_excel(data, excel_filename="OS第一次作业汇总.xlsx"):
    # 转换为 pandas DataFrame
    df = pd.DataFrame(data)

    # 设置表头的映射
    df.columns = ['学号', '姓名', '机器检查得分']

    # 保存为Excel文件
    df.to_excel(excel_filename, index=False)

    # 使用openpyxl加载Excel，调整样式
    wb = load_workbook(excel_filename)
    ws = wb.active

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)  # 加粗
        cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平和垂直居中

    # 设置数据单元格居中
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽，考虑表头和内容的宽度
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列的字母
        for i, cell in enumerate(col):
            try:
                # 对比表头和内容的长度，选择较长的
                cell_length = len(str(cell.value))
                if i == 0:  # 如果是表头
                    max_length = cell_length
                else:
                    max_length = max(max_length, cell_length)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 设置列宽的比例
        ws.column_dimensions[column].width = adjusted_width

    # 固定设置第三列宽度为20
    ws.column_dimensions[get_column_letter(3)].width = 20

    # 自动调整行高
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20  # 设置默认行高为20

    # 保存修改后的Excel
    wb.save(excel_filename)

    print(f"Excel文件已成功生成: {excel_filename}")

if __name__ == '__main__':
    generate_excel(data)
