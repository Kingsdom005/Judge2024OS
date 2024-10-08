from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import yaml
import re
import os
import sys

print("目录初始化中……")

def generate_folder(folder_path):
    # 判断文件夹是否存在
    if not os.path.exists(folder_path):
        # 如果不存在，则创建文件夹
        os.makedirs(folder_path)

generate_folder("./data")
generate_folder("./data/second")
generate_folder("./excel")
generate_folder("./reports")
generate_folder("./reports/second")
generate_folder("./pytest")

print("目录初始化完成！")

# 读取markdown文档
directory = './data/second/'
md_files = [file for file in os.listdir(directory) if file.endswith('.md')]
md_first_files = [item for item in md_files if '第二次' in item]
excel_data = []
default_read_md_file = ""
# print(md_first_files)

def get_structure_info(file_name, prefix="./data/second/"):
    res = {
        "filename": '',
        "l1_title": '',
        "l2_titles": []
    }
    # 读取Markdown文件
    with open(prefix + file_name, 'r', encoding='utf-8') as file:
        global markdown_content
        markdown_content = file.read()

    # 匹配所有一级标题和二级标题
    level1_titles = re.findall(r'^(# .+)', markdown_content, re.MULTILINE)
    level2_titles = re.findall(r'^(## .+)', markdown_content, re.MULTILINE)

    # 输出文件名
    # print(f"文件名: {file_name}\n")
    res["filename"] = file_name

    # 输出一级标题
    # print("一级标题:")
    for title in level1_titles:
        # print(title[2:])  # 去掉"# "
        res["l1_title"] = title[2:]

    # 输出二级标题
    # print("\n二级标题:")
    for title in level2_titles:
        # print(title[3:])  # 去掉"## "
        res["l2_titles"].append(title[3:])

    return res

# 处理代码，关键词检索
def find_targets(source, targets):
    # 拆分 source 为多行
    lines = source.splitlines()

    # 结果列表
    res = []

    # 遍历目标字符串
    for target in targets:
        found = False
        # 遍历每一行，检测是否包含目标字符串
        for i, line in enumerate(lines, start=1):  # 行号从1开始
            if target in line:
                res.append(i)  # 保存行号
                found = True
                break
        if not found:
            res.append(-1)  # 没有找到该目标时返回 -1

    return res

def format_number(num):
    # 判断是否是整数
    if num == int(num):
        return f"{int(num)}"
    else:
        return f"{num:.2f}"

def generate_excel(data, excel_filename="OS第二次作业汇总.xlsx"):
    # 转换为 pandas DataFrame
    df = pd.DataFrame(data)

    # 设置表头的映射
    df.columns = ['学号', '姓名', '机器检查得分']

    # 保存为Excel文件
    excel_filename = "./excel/" + excel_filename
    df.to_excel(excel_filename, index=False)

    # 使用openpyxl加载Excel，调整样式
    wb = load_workbook(excel_filename)
    ws = wb.active

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)  # 加粗
        cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平和垂直居中

    # 设置数据单元格居中
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽，考虑表头和内容的宽度
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列的字母
        for i, cell in enumerate(col):
            try:
                # 对比表头和内容的长度，选择较长的
                cell_length = len(str(cell.value))
                if i == 0:  # 如果是表头
                    max_length = cell_length
                else:
                    max_length = max(max_length, cell_length)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 设置列宽的比例
        ws.column_dimensions[column].width = adjusted_width

    # 固定设置第三列宽度为20
    ws.column_dimensions[get_column_letter(3)].width = 20

    ws.column_dimensions[get_column_letter(2)].width = 15

    # 自动调整行高
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 20  # 设置默认行高为20

    # 保存修改后的Excel
    wb.save(excel_filename)

    print(f"Excel文件已成功生成: {excel_filename}")


if __name__ == '__main__':

    if len(md_first_files) == 0:
        print("没有检测到第二次作业相关文档！\n请检查文件命名是否为：学号-姓名-操作系统第二次作业.md")
        sys.exit()

    for first_name in md_first_files:
        print("检索markdown文档中……")
        default_read_md_file = first_name

        print(f"检索成功，检索到{default_read_md_file}")
        print("正在解析文档结构……")

        markdown_content = ""

        structure = get_structure_info(default_read_md_file)

        # def check_structure():
        print("解析文档结构成功")
        print(structure)

        # read config yml file
        print("读取配置文件中……")
        with open('config2.yml', 'r', encoding="utf8") as file:
            config = yaml.safe_load(file)

        print(config)
        print("读取成功")

        check_structure = {
            "filename": False,
            "structure_l1": False,
            "structure_l2": [False for _ in range(5)]
        }

        pattern = r'^\d{10}-\w+-操作系统第二次作业\.md$'
        if re.match(pattern, structure["filename"]):
            check_structure["filename"] = True

        if structure["l1_title"] == config["judgeConfig"]["structure"]["l1_title"]:
            check_structure["structure_l1"] = True

        for i in range(len(structure["l2_titles"])):
            if structure["l2_titles"][i] == config["judgeConfig"]["structure"]["l2_titles"][i]:
                check_structure["structure_l2"][i] = True

        print("检查markdown结构中……")
        print(check_structure)
        structure_check_flag = True
        if not check_structure["filename"]:
            print("请检查文件名！！")
            structure_check_flag = False
            # sys.exit()
        if not check_structure["structure_l1"]:
            print("请检查一级标题！！")
            structure_check_flag = False
            # sys.exit()
        for l2_subtitle in check_structure["structure_l2"]:
            if not l2_subtitle:
                print("请检查二级标题！！")
                structure_check_flag = False
                # sys.exit()

        print("检测结束，结构无误.")

        print("开始判题")

        # 读取Markdown文件
        # 使用正则表达式提取所需文本
        patterns = [
            r'## 1\.实验内容\n(.*?)## 2\.实验思路',
            r'## 2\.实验思路\n(.*?)## 3\.实验源码',
            r'## 3\.实验源码\n(.*?)## 4\.实验结果',
            r'## 4\.实验结果\n(.*?)## 5\.实验总结与反思',
            r'## 5\.实验总结与反思\n(.*)',
        ]

        total_score = 0

        limitations = config["judgeConfig"]["limitations"]
        scores = config["judgeConfig"]["scores"]

        # 表格内容
        data = [
            ["实验内容", 0.5, 0],
            ["实验思路", 1, 0],
            ["实验源码", 4, 0],
            ["实验结果", 3, 0],
            ["实验总结与反思", 1, 0],
            ["文件结构和名称", 0.5, 0],
        ]

        if structure_check_flag == True:
            total_score += 0.5
            data[-1][2] = data[-1][1]

        for index, ppp in enumerate(patterns):

            pattern = re.compile(ppp, re.DOTALL)
            match = pattern.search(markdown_content)

            # print('\n\n\nmatch:\n',match, "\n\n\n")

            if match:
                content_between = match.group(1).strip()
                print(f"所属章节：{data[index][0]}")
                print("提取的文本:")
                print(content_between)
                if data[index][0] == "实验结果":
                    # 需要进一步拆分实验结果
                    tmp_res_score = 0
                    per_score = float(scores[index]["score"]) / 5
                    # 使用正则表达式提取“测试结果”和“原理分析”的内容
                    pattern = r"测试结果：\s*(.*?)\n.*?原理分析：\s*(.*?)\n" # 可能匹配不到最后一个
                    matches = re.findall(pattern, content_between, re.DOTALL)
                    print("matches:", matches)

                    # 输出结果
                    for i, (test_result, analysis) in enumerate(matches, start=1):
                        res = re.findall(r'\d+', test_result)[0]
                        print(f"测试内容{i}：{res}")
                        if i == 1:
                            if int(res) < 200000 and int(res) > 1000: 
                                tmp_res_score += per_score
                                print(f"测试结果{i}：通过")
                            else:
                                print(f"测试结果{i}：通过")
                        elif i == 2 or i == 4 :
                            if int(res) == 200000:
                                tmp_res_score += per_score
                                print(f"测试结果{i}：通过")
                            else:
                                print(f"测试结果{i}：不通过")
                        elif i == 3:
                            if int(res) <= 200000 and int(res) >= 190000:
                                tmp_res_score += per_score
                                print(f"测试结果{i}：通过")
                            else:
                                print(f"测试结果{i}：不通过")
                                
                    final_test_res = content_between.split("生产者消费者问题")[-1].strip().split("\n")[0].split("测试结果：")[-1].strip()
                    final_analysis_res = content_between.split("生产者消费者问题")[-1].strip().split("\n")[-1].split("原理分析：")[-1].strip()
                    print(f"测试结内容5：{final_test_res}")
                    # print(final_test_res, "\n" ,final_analysis_res)
                    if len(final_test_res) > 10 and len(final_analysis_res) > 10:
                        tmp_res_score += per_score
                        print(f"测试结果5：通过")
                    else:
                        print(f"测试结果5：不通过")

                    total_score += tmp_res_score
                    data[index][2] = tmp_res_score

                elif len(content_between) >= float(limitations[index]["min_len"]):
                    total_score += float(scores[index]["score"])
                    data[index][2] = float(scores[index]["score"])
            else:
                print("未找到指定段落之间的内容。")

        print("机器检查最终得分：", total_score)


        # generate html
        # HTML模板
        html_template = """
        <!DOCTYPE html>
        <html lang="zh">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>实验报告</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    margin: 0;
                    padding: 20px;
                }}
                h1 {{
                    text-align: center;
                    color: #333;
                }}
                table {{
                    width: 60%;
                    margin: 0 auto;
                    border-collapse: collapse;
                    background-color: #fff;
                    box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
                }}
                th, td {{
                    border: 1px solid #ddd;
                    padding: 12px;
                    text-align: center;
                }}
                th {{
                    background-color: #4CAF50;
                    color: white;
                }}
                tr:nth-child(even) {{
                    background-color: #f2f2f2;
                }}
                tr:hover {{
                    background-color: #ddd;
                }}
                .note {{
                    margin-top: 20px;
                    font-size: 14px;
                    color: #777;
                    width: 60%; /* 与表格同宽 */
                    margin-left: auto;
                    margin-right: auto; /* 使其居中 */
                }}
            </style>
        </head>
        <body>
            <h1>实验报告</h1>
            <table>
                <tr>
                    <th>测试功能点</th>
                    <th>分值</th>
                    <th>实际得分</th>
                </tr>
                {rows}
            </table>
            <div class="note">
                *说明：本机器检测结果并不代表最终实际得分。
            </div>
        </body>
        </html>
        """

        # # 表格内容
        # data = [
        #     ("文件结构和名称", 1, 1),
        #     ("实验思路", 1, 1),
        #     ("实验源码", 4, 1),
        #     ("实验结果", 2, 2),
        #     ("实验总结与反思", 1, 1),
        # ]

        # 生成表格行
        rows = ""
        for item in data:
            rows += f"<tr><td>{item[0]}</td><td>{item[1]}</td><td>{format_number(item[2])}</td></tr>\n"
        rows += f"<tr><td>{'总分'}</td><td>{10}</td><td>{format_number(total_score)}</td></tr>\n"
        # 生成完整HTML
        html_content = html_template.format(rows=rows)

        file_info = {
            "id" : default_read_md_file.split("-")[0],
            "name" : default_read_md_file.split("-")[1],
        }

        # 保存html文件
        with open(f'./reports/second/{file_info["id"]}-{file_info["name"]}-report.html', 'w', encoding='utf-8') as file:
            file.write(html_content)

        print("HTML报告已生成！")

        excel_data.append({'id': file_info["id"], 'name': file_info["name"], 'score': total_score})

    # generate excel
    generate_excel(excel_data)
